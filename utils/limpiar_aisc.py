"""
limpiar_aisc.py
---------------
Limpia el catálogo AISC Shapes Database v14.1 y genera un Excel listo
para usar en el optimizador A* estructural (v3).
 
Uso:
    python limpiar_aisc.py
    python limpiar_aisc.py --input mi_archivo.xls --tipo W,HP --output catalogo.xlsx
 
Salida:
    AISC_Catalogo_Limpio.xlsx  →  una pestaña por tipo de perfil seleccionado
"""


import argparse
import sys
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_DEFAULT  = "AISC_Shapes_Database_v14_1.xls"
OUTPUT_DEFAULT = "AISC_Catalogo_Limpio.xlsx"
SHEET_RAW      = "Shapes_AISC_V14"
TIPOS_DEFAULT  = ["W"]          # filtrar solo W-shapes por defecto
 
# Columnas a conservar y sus alias limpios
COLS_MAP = {
    "AISC_Manual_Label" : "nombre",
    "W"                 : "peso_lbft",      # lb/ft
    "A"                 : "area_in2",       # in²
    "d"                 : "d_in",           # altura total, in
    "bf"                : "bf_in",          # ancho de ala, in
    "tw"                : "tw_in",          # espesor alma, in
    "tf"                : "tf_in",          # espesor ala, in
    "Ix"                : "Ix_in4",         # inercia eje fuerte, in⁴
    "Sx"                : "Sx_in3",         # módulo elástico fuerte, in³
    "Zx"                : "Zx_in3",         # módulo plástico fuerte, in³
    "rx"                : "rx_in",          # radio de giro fuerte, in
    "Iy"                : "Iy_in4",         # inercia eje débil, in⁴
    "Sy"                : "Sy_in3",         # módulo elástico débil, in³
    "Zy"                : "Zy_in3",         # módulo plástico débil, in³
    "ry"                : "ry_in",          # radio de giro débil, in
    "J"                 : "J_in4",          # constante torsional, in⁴
    "Cw"                : "Cw_in6",         # constante de alabeo, in⁶
    "rts"               : "rts_in",         # radio de giro efectivo LTB, in
    "ho"                : "ho_in",          # dist. centros de alas, in
    "bf/2tf"            : "bf_2tf",         # relación compacidad ala
    "h/tw"              : "h_tw",           # relación compacidad alma
}
 
# Columnas numéricas (todas excepto nombre)
COLS_NUM = [v for v in COLS_MAP.values() if v != "nombre"]
 
# Factores de conversión a SI (para hoja extra)
# 1 in  = 0.0254 m  |  1 lb/ft = 1.48816 kg/m
IN_TO_M   = 0.0254
IN2_TO_M2 = IN_TO_M ** 2
IN3_TO_M3 = IN_TO_M ** 3
IN4_TO_M4 = IN_TO_M ** 4
IN6_TO_M6 = IN_TO_M ** 6
LBFT_TO_KGM = 1.48816
 
FACTORES_SI = {
    "peso_lbft" : ("peso_kgm",  LBFT_TO_KGM),
    "area_in2"  : ("area_m2",   IN2_TO_M2),
    "d_in"      : ("d_m",       IN_TO_M),
    "bf_in"     : ("bf_m",      IN_TO_M),
    "tw_in"     : ("tw_m",      IN_TO_M),
    "tf_in"     : ("tf_m",      IN_TO_M),
    "Ix_in4"    : ("Ix_m4",     IN4_TO_M4),
    "Sx_in3"    : ("Sx_m3",     IN3_TO_M3),
    "Zx_in3"    : ("Zx_m3",     IN3_TO_M3),
    "rx_in"     : ("rx_m",      IN_TO_M),
    "Iy_in4"    : ("Iy_m4",     IN4_TO_M4),
    "Sy_in3"    : ("Sy_m3",     IN3_TO_M3),
    "Zy_in3"    : ("Zy_m3",     IN3_TO_M3),
    "ry_in"     : ("ry_m",      IN_TO_M),
    "J_in4"     : ("J_m4",      IN4_TO_M4),
    "Cw_in6"    : ("Cw_m6",     IN6_TO_M6),
    "rts_in"    : ("rts_m",     IN_TO_M),
    "ho_in"     : ("ho_m",      IN_TO_M),
    "bf_2tf"    : ("bf_2tf",    1.0),       # adimensional
    "h_tw"      : ("h_tw",      1.0),       # adimensional
}
 
# ─── Estilos Excel ─────────────────────────────────────────────────────────────
 
COLOR_HEADER   = "1F3864"   # azul oscuro
COLOR_SUBHEAD  = "2E75B6"   # azul medio
COLOR_ALT      = "D6E4F0"   # azul claro (filas alternas)
COLOR_WHITE    = "FFFFFF"
FONT_WHITE     = Font(name="Arial", bold=True, color="FFFFFF", size=10)
FONT_NORMAL    = Font(name="Arial", size=10)
FONT_BOLD      = Font(name="Arial", bold=True, size=10)
ALIGN_CENTER   = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT     = Alignment(horizontal="left",   vertical="center")
ALIGN_RIGHT    = Alignment(horizontal="right",  vertical="center")
 
def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
 
# ─── Funciones ─────────────────────────────────────────────────────────────────
 
def cargar_raw(path: str) -> pd.DataFrame:
    """Lee la hoja principal del AISC omitiendo la fila de unidades SI."""
    df = pd.read_excel(path, sheet_name=SHEET_RAW, header=0, skiprows=[1])
    return df
 
 
def limpiar_tipo(df: pd.DataFrame, tipo: str) -> pd.DataFrame:
    """Filtra un tipo, selecciona columnas, limpia y convierte a numérico."""
    sub = df[df["Type"] == tipo][list(COLS_MAP.keys())].copy()
    sub = sub.rename(columns=COLS_MAP)
 
    # Reemplazar el guión AISC "–" por NaN y convertir columnas numéricas
    for col in COLS_NUM:
        sub[col] = pd.to_numeric(
            sub[col].replace("–", np.nan), errors="coerce"
        )
 
    # Eliminar filas donde TODAS las columnas numéricas sean NaN
    sub = sub.dropna(how="all", subset=COLS_NUM)
 
    # Ordenar por peso ascendente (más liviano primero, igual que A*)
    sub = sub.sort_values("peso_lbft", ascending=True).reset_index(drop=True)
 
    return sub
 
 
def agregar_si(df_imperial: pd.DataFrame) -> pd.DataFrame:
    """Agrega columna nombre + todas las columnas convertidas a SI."""
    si = pd.DataFrame()
    si["nombre"] = df_imperial["nombre"]
    for col_imp, (col_si, factor) in FACTORES_SI.items():
        if col_imp in df_imperial.columns:
            si[col_si] = (df_imperial[col_imp] * factor).round(8)
    return si
 
 
def reporte_limpieza(df_original: pd.DataFrame, df_limpio: pd.DataFrame, tipo: str):
    orig  = len(df_original[df_original["Type"] == tipo])
    final = len(df_limpio)
    nulls = df_limpio[COLS_NUM].isnull().sum()
    cols_con_nulos = nulls[nulls > 0]
 
    print(f"\n  [{tipo}] {orig} perfiles → {final} conservados "
          f"({orig - final} eliminados)")
    if not cols_con_nulos.empty:
        print(f"  Columnas con NaN restantes (no eliminan filas):")
        for col, n in cols_con_nulos.items():
            print(f"    • {col}: {n} nulos")
    else:
        print(f"  Sin NaN en columnas numéricas ✓")
 
 
def escribir_hoja(wb, nombre_hoja: str, df_imperial: pd.DataFrame, df_si: pd.DataFrame):
    """Escribe una hoja con subencabezados Imperial / SI y formato profesional."""
    ws = wb.create_sheet(title=nombre_hoja)
 
    # ── Fila 1: título de la hoja ──────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(df_imperial.columns) + len(df_si.columns) - 1 + 1)
    title_cell = ws.cell(row=1, column=1,
                         value=f"Perfiles {nombre_hoja} — AISC 360 v14.1")
    title_cell.font    = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    title_cell.fill    = PatternFill("solid", fgColor=COLOR_HEADER)
    title_cell.alignment = ALIGN_CENTER
 
    # ── Fila 2: subencabezados de sección ─────────────────────────────────
    n_imp = len(df_imperial.columns)
    n_si  = len(df_si.columns) - 1  # excluye 'nombre' duplicado
 
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_imp)
    imp_cell = ws.cell(row=2, column=1, value="UNIDADES IMPERIALES  (in, lb/ft, in², in³, in⁴, in⁶)")
    imp_cell.font    = FONT_WHITE
    imp_cell.fill    = PatternFill("solid", fgColor=COLOR_SUBHEAD)
    imp_cell.alignment = ALIGN_CENTER
 
    ws.merge_cells(start_row=2, start_column=n_imp + 1,
                   end_row=2, end_column=n_imp + n_si)
    si_cell = ws.cell(row=2, column=n_imp + 1,
                      value="UNIDADES SI  (m, kg/m, m², m³, m⁴, m⁶)")
    si_cell.font    = FONT_WHITE
    si_cell.fill    = PatternFill("solid", fgColor="1A5276")
    si_cell.alignment = ALIGN_CENTER
 
    # ── Fila 3: nombres de columnas ────────────────────────────────────────
    all_cols = list(df_imperial.columns)
    si_cols  = [c for c in df_si.columns if c != "nombre"]
    all_cols += si_cols
 
    for col_idx, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font      = FONT_WHITE
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = ALIGN_CENTER
        cell.border    = thin_border()
 
    # ── Filas de datos ─────────────────────────────────────────────────────
    for row_idx, (_, row_imp) in enumerate(df_imperial.iterrows(), start=4):
        fill_color = COLOR_ALT if row_idx % 2 == 0 else COLOR_WHITE
        fill       = PatternFill("solid", fgColor=fill_color)
 
        # Columnas imperiales
        for col_idx, val in enumerate(row_imp, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = FONT_NORMAL
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = ALIGN_LEFT if col_idx == 1 else ALIGN_RIGHT
 
        # Columnas SI (misma fila, sin 'nombre')
        row_si = df_si.iloc[row_idx - 4]
        for col_offset, col_name in enumerate(si_cols, start=1):
            val  = row_si[col_name]
            cell = ws.cell(row=row_idx, column=n_imp + col_offset, value=val)
            cell.font      = FONT_NORMAL
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = ALIGN_RIGHT
 
    # ── Anchos de columna ─────────────────────────────────────────────────
    for col_idx, col_name in enumerate(all_cols, start=1):
        width = max(len(col_name) + 2, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
 
    # Fijar filas de encabezado
    ws.freeze_panes = "B4"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
 
 
def escribir_hoja_info(wb, resumen: list[dict]):
    """Hoja README con descripción de cada columna."""
    ws = wb.create_sheet(title="README")
 
    titulo = ws.cell(row=1, column=1, value="Diccionario de columnas — AISC Catalog Cleaner")
    titulo.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    titulo.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    titulo.alignment = ALIGN_CENTER
    ws.merge_cells("A1:D1")
 
    headers = ["Columna imperial", "Columna SI", "Descripción", "Unidades SI"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font  = FONT_WHITE
        cell.fill  = PatternFill("solid", fgColor=COLOR_SUBHEAD)
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border()
 
    descripciones = [
        ("nombre",    "nombre",   "Designación AISC del perfil",              "—"),
        ("peso_lbft", "peso_kgm", "Peso lineal",                              "kg/m"),
        ("area_in2",  "area_m2",  "Área de la sección transversal",           "m²"),
        ("d_in",      "d_m",      "Altura total del perfil",                  "m"),
        ("bf_in",     "bf_m",     "Ancho de ala",                             "m"),
        ("tw_in",     "tw_m",     "Espesor del alma",                         "m"),
        ("tf_in",     "tf_m",     "Espesor de ala",                           "m"),
        ("Ix_in4",    "Ix_m4",    "Momento de inercia eje fuerte (x-x)",      "m⁴"),
        ("Sx_in3",    "Sx_m3",    "Módulo elástico eje fuerte",               "m³"),
        ("Zx_in3",    "Zx_m3",    "Módulo plástico eje fuerte → φMn",        "m³"),
        ("rx_in",     "rx_m",     "Radio de giro eje fuerte → pandeo",        "m"),
        ("Iy_in4",    "Iy_m4",    "Momento de inercia eje débil (y-y)",       "m⁴"),
        ("Sy_in3",    "Sy_m3",    "Módulo elástico eje débil",                "m³"),
        ("Zy_in3",    "Zy_m3",    "Módulo plástico eje débil",                "m³"),
        ("ry_in",     "ry_m",     "Radio de giro eje débil → pandeo",         "m"),
        ("J_in4",     "J_m4",     "Constante torsional de St. Venant",        "m⁴"),
        ("Cw_in6",    "Cw_m6",    "Constante de alabeo (warping)",            "m⁶"),
        ("rts_in",    "rts_m",    "Radio de giro efectivo → LTB (pandeo lat.)","m"),
        ("ho_in",     "ho_m",     "Distancia entre centros de alas → LTB",    "m"),
        ("bf_2tf",    "bf_2tf",   "Relación bf/2tf — compacidad de ala",       "adim."),
        ("h_tw",      "h_tw",     "Relación h/tw — compacidad de alma",        "adim."),
    ]
 
    fill_alt = PatternFill("solid", fgColor=COLOR_ALT)
    fill_wh  = PatternFill("solid", fgColor=COLOR_WHITE)
 
    for ri, (ci, csi, desc, uni) in enumerate(descripciones, start=3):
        fill = fill_alt if ri % 2 == 0 else fill_wh
        for ci_idx, val in enumerate([ci, csi, desc, uni], start=1):
            cell = ws.cell(row=ri, column=ci_idx, value=val)
            cell.font  = FONT_NORMAL
            cell.fill  = fill
            cell.border = thin_border()
            cell.alignment = ALIGN_LEFT
 
    # Nota de conversión
    nota_row = len(descripciones) + 4
    nota = ws.cell(row=nota_row, column=1,
                   value="Factores de conversión: 1 in = 0.0254 m  |  1 lb/ft = 1.48816 kg/m  |  E acero = 200 GPa = 2.0×10⁸ kN/m²")
    nota.font = Font(name="Arial", italic=True, size=9, color="555555")
    ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=4)
 
    for col_idx, width in enumerate([18, 14, 42, 10], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
 
    ws.row_dimensions[1].height = 22
 
    # Resumen de conteo
    res_row = nota_row + 2
    ws.cell(row=res_row, column=1, value="Resumen por tipo:").font = FONT_BOLD
    for i, r in enumerate(resumen, start=res_row + 1):
        ws.cell(row=i, column=1, value=r["tipo"]).font = FONT_NORMAL
        ws.cell(row=i, column=2, value=r["count"]).font = FONT_NORMAL
 
 
# ─── Main ──────────────────────────────────────────────────────────────────────
 
def main():
    parser = argparse.ArgumentParser(description="Limpieza catálogo AISC")
    parser.add_argument("--input",  default=INPUT_DEFAULT,  help="Archivo .xls de entrada")
    parser.add_argument("--output", default=OUTPUT_DEFAULT, help="Archivo .xlsx de salida")
    parser.add_argument("--tipo",   default=",".join(TIPOS_DEFAULT),
                        help="Tipos de perfil separados por coma (p.ej. W,HP,C)")
    args = parser.parse_args()
 
    tipos   = [t.strip().upper() for t in args.tipo.split(",")]
    in_path = Path(args.input)
    out_path = Path(args.output)
 
    print(f"\n{'='*55}")
    print(f"  AISC Catalog Cleaner")
    print(f"  Entrada : {in_path}")
    print(f"  Salida  : {out_path}")
    print(f"  Tipos   : {tipos}")
    print(f"{'='*55}")
 
    if not in_path.exists():
        sys.exit(f"❌  No se encontró el archivo: {in_path}")
 
    print("\n▶ Cargando datos crudos...")
    df_raw = cargar_raw(str(in_path))
    print(f"  {len(df_raw)} filas totales, {df_raw['Type'].nunique()} tipos de perfil")
 
    # Crear workbook y eliminar hoja por defecto
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
 
    resumen = []
    for tipo in tipos:
        if tipo not in df_raw["Type"].values:
            print(f"\n  ⚠ Tipo '{tipo}' no encontrado en el catálogo — omitido")
            continue
 
        print(f"\n▶ Procesando tipo: {tipo}")
        df_imp = limpiar_tipo(df_raw, tipo)
        df_si  = agregar_si(df_imp)
        reporte_limpieza(df_raw, df_imp, tipo)
 
        escribir_hoja(wb, tipo, df_imp, df_si)
        resumen.append({"tipo": tipo, "count": len(df_imp)})
 
    if not resumen:
        sys.exit("❌  Ningún tipo procesado. Revisa el parámetro --tipo.")
 
    escribir_hoja_info(wb, resumen)
 
    # Mover README al principio
    wb.move_sheet("README", offset=-len(wb.sheetnames) + 1)
 
    wb.save(str(out_path))
    print(f"\n{'='*55}")
    print(f"✅  Archivo guardado: {out_path}")
    for r in resumen:
        print(f"   • {r['tipo']}: {r['count']} perfiles")
    print(f"{'='*55}\n")
 
 
if __name__ == "__main__":
    main()